import openpyxl

FILENAME = "telephone_book.xlsx"

# EXCEL FUNCTION HANDLING
def init_excel():
    try:
        wb = openpyxl.load_workbook(FILENAME)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Phones"
        ws.append(["Phone Number"])
        wb.save(FILENAME)

def load_book():
    wb = openpyxl.load_workbook(FILENAME)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            data.append(row[0])
    return data

def save_book():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Phones"
    ws.append(["Phone Number"])
    for number in book:
        ws.append([number])
    wb.save(FILENAME)
    
#SCRIPT LOGIC

book = []

def add():
    addition = input("Enter phone number: ")
    if addition not in book:
        book.append(addition)
        save_book()
        print("Phone added!")
    else:
        print("Phone already exists!")

def delete():
    if book:
        deletion = input("Delete phone number: ")
        if deletion in book:
            book.remove(deletion)
            save_book()
            print("Phone number removed!")
        else:
            print("Phone number doesn't exist")
    else:
        print("Book is empty")

def view():
    if book:
        print("All phone numbers:")
        for number in book:
            print(" -", number)
    else:
        print("Book is empty")

def main():
    init_excel()
    global book
    book = load_book()

    while True:
        selection = input("1. Add   2. Remove   3. View   4. Exit: ")
        if selection == "1":
            add()
        elif selection == "2":
            delete()
        elif selection == "3":
            view()
        elif selection == "4":
            print("Exiting...")
            break
        else:
            print("Invalid choice, try again")

main()
